# -*- coding: utf-8 -*-
"""
Load in time data from Rattlesnake runs

Copyright 2022 National Technology & Engineering Solutions of Sandia,
LLC (NTESS). Under the terms of Contract DE-NA0003525 with NTESS, the U.S.
Government retains certain rights in this software.

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

import netCDF4 as nc4
import numpy as np
from ..core.sdynpy_coordinate import (coordinate_array, outer_product,
                                      CoordinateArray, _string_map)
from ..core.sdynpy_data import data_array, FunctionTypes
from ..core.sdynpy_system import System
import pandas as pd
import sys
import openpyxl as opxl
import os
import warnings


def read_rattlesnake_output(file, coordinate_override_column=None, read_only_indices=None,
                            read_variable='time_data', abscissa_start = None,
                            abscissa_stop = None, downsample = None):
    """
    Reads in a Rattlesnake data file and returns the time history array as well
    as the channel table

    Parameters
    ----------
    file : str or nc4.Dataset
        Path to the file to read in or an already open
    coordinate_override_column : str, optional
        Specify a channel table column name to extract coordinate information from.
        If not specified, then it will be assembled from node id and directions.
    read_only_indices : slice or iterable, optional
        A valid indexing operation to select which channel indices to read
    read_variable : str, optional
        The time variable from the Rattlesnake file to read.  These will 
        generally be time_data, time_data_1, time_data_2, etc. depending on
        how many streams exist in the file.  The default is 'time_data'.
    abscissa_start : float, optional
        Data will not be extracted for abscissa values less than this value
    abscissa_stop : float, optional
        Data will not be extracted for abscissa values greater than this value
    downsample : int, optional
        A step size to use to downsample the dataset when reading

    Returns
    -------
    data_array : TimeHistoryArray
        Time history data in the Rattlesnake output file
    channel_table : DataFrame
        Pandas Dataframe containing the channel table information

    """
    if isinstance(file, str):
        ds = nc4.Dataset(file, 'r')
    elif isinstance(file, nc4.Dataset):
        ds = file
    if read_only_indices is None:
        read_only_indices = slice(None)
    if abscissa_start is None:
        start_index = None
    else:
        start_index = int(np.ceil(abscissa_start * ds.sample_rate))
    if abscissa_stop is None:
        stop_index = None
    else:
        stop_index = int(np.ceil(abscissa_stop * ds.sample_rate))
    abscissa_slice = slice(start_index, stop_index, downsample)
    output_data = np.array(ds[read_variable][:,abscissa_slice][read_only_indices])
    abscissa = np.arange(0 if start_index is None else start_index,
                         ds[read_variable].shape[-1] if stop_index is None else stop_index,
                         1 if downsample is None else downsample) / ds.sample_rate
    if coordinate_override_column is None:
        nodes = [int(''.join(char for char in node if char in '0123456789'))
                 for node in ds['channels']['node_number'][...][read_only_indices]]
        directions = np.array(ds['channels']['node_direction'][...][read_only_indices], dtype='<U3')
        coordinates = coordinate_array(nodes, directions)[:, np.newaxis]
    else:
        coordinates = coordinate_array(string_array=ds['channels'][coordinate_override_column][read_only_indices])[
            :, np.newaxis]
    array = {name: np.array(variable[:]) for name, variable in ds['channels'].variables.items()}
    channel_table = pd.DataFrame(array)
    comment1 = np.char.add(np.char.add(np.array(ds['channels']['channel_type'][...][read_only_indices], dtype='<U80'),
                                       np.array(' :: ')),
                           np.array(ds['channels']['unit'][...][read_only_indices], dtype='<U80'))
    comment2 = np.char.add(np.char.add(np.array(ds['channels']['physical_device'][...][read_only_indices], dtype='<U80'),
                                       np.array(' :: ')),
                           np.array(ds['channels']['physical_channel'][...][read_only_indices], dtype='<U80'))
    comment3 = np.char.add(np.char.add(np.array(ds['channels']['feedback_device'][...][read_only_indices], dtype='<U80'),
                                       np.array(' :: ')),
                           np.array(ds['channels']['feedback_channel'][...][read_only_indices], dtype='<U80'))
    comment4 = np.array(ds['channels']['comment'][...][read_only_indices], dtype='<U80')
    comment5 = np.array(ds['channels']['make'][...][read_only_indices], dtype='<U80')
    for key in ('model', 'serial_number', 'triax_dof'):
        comment5 = np.char.add(comment5, np.array(' '))
        comment5 = np.char.add(comment5, np.array(ds['channels'][key][...][read_only_indices], dtype='<U80'))
    time_data = data_array(FunctionTypes.TIME_RESPONSE,
                           abscissa,
                           output_data,
                           coordinates,
                           comment1,
                           comment2,
                           comment3,
                           comment4,
                           comment5)
    if isinstance(file, str):
        ds.close()
    return time_data, channel_table


def read_system_id_data(file):
    if isinstance(file, str):
        file = np.load(file)
    df = file['sysid_frequency_spacing']
    if np.isnan(file['response_transformation_matrix']):
        try:
            response_dofs = coordinate_array(
                [int(v) for v in file['channel_node_number'][file['response_indices']]],
                file['channel_node_direction'][file['response_indices']])
        except Exception:
            response_dofs = coordinate_array(file['response_indices']+1, 0)
    else:
        response_dofs = coordinate_array(np.arange(file['response_transformation_matrix'].shape[0])+1, 0)
    if np.isnan(file['reference_transformation_matrix']):
        try:
            reference_dofs = coordinate_array(
                [int(v) for v in file['channel_node_number'][file['reference_indices']]],
                file['channel_node_direction'][file['reference_indices']])
        except Exception:
            reference_dofs = coordinate_array(file['reference_indices']+1, 0)
    else:
        reference_dofs = coordinate_array(np.arange(file['reference_transformation_matrix'].shape[0])+1, 0)
    ordinate = np.moveaxis(file['frf_data'], 0, -1)
    frfs = data_array(FunctionTypes.FREQUENCY_RESPONSE_FUNCTION,
                      df*np.arange(ordinate.shape[-1]), ordinate,
                      outer_product(response_dofs, reference_dofs))
    ordinate = np.moveaxis(file['response_cpsd'], 0, -1)
    response_cpsd = data_array(FunctionTypes.POWER_SPECTRAL_DENSITY,
                               df*np.arange(ordinate.shape[-1]), ordinate,
                               outer_product(response_dofs, response_dofs))
    ordinate = np.moveaxis(file['reference_cpsd'], 0, -1)
    reference_cpsd = data_array(FunctionTypes.POWER_SPECTRAL_DENSITY,
                                df*np.arange(ordinate.shape[-1]), ordinate,
                                outer_product(reference_dofs, reference_dofs))
    ordinate = np.moveaxis(file['response_noise_cpsd'], 0, -1)
    response_noise_cpsd = data_array(FunctionTypes.POWER_SPECTRAL_DENSITY,
                                     df*np.arange(ordinate.shape[-1]), ordinate,
                                     outer_product(response_dofs, response_dofs))
    ordinate = np.moveaxis(file['reference_noise_cpsd'], 0, -1)
    reference_noise_cpsd = data_array(FunctionTypes.POWER_SPECTRAL_DENSITY,
                                      df*np.arange(ordinate.shape[-1]), ordinate,
                                      outer_product(reference_dofs, reference_dofs))
    ordinate = np.moveaxis(file['coherence'], 0, -1)
    coherence = data_array(FunctionTypes.MULTIPLE_COHERENCE,
                           df*np.arange(ordinate.shape[-1]), ordinate,
                           outer_product(response_dofs))
    return frfs, response_cpsd, reference_cpsd, response_noise_cpsd, reference_noise_cpsd, coherence

def read_system_id_nc4(file, coordinate_override_column=None):
    if isinstance(file,str):
        ds = nc4.Dataset(file,'r')
    elif isinstance(file,nc4.Dataset):
        ds = file
    
    environment = [group for group in ds.groups if not group == 'channels'][0]

    # Get the channels in the group
    if coordinate_override_column is None:
        nodes = [int(''.join(char for char in node if char in '0123456789'))
                 for node in ds['channels']['node_number']]
        directions = np.array(ds['channels']['node_direction'][:], dtype='<U3')
        coordinates = coordinate_array(nodes, directions)
    else:
        coordinates = coordinate_array(string_array=ds['channels'][coordinate_override_column])
    drives = ds['channels']['feedback_device'][:] != ''

    # Cull down to just those in the environment
    environment_index = np.where(ds['environment_names'][:] == environment)[0][0]
    environment_channels = ds['environment_active_channels'][:, environment_index].astype(bool)

    drives = drives[environment_channels]
    coordinates = coordinates[environment_channels]

    control_indices = ds[environment]['control_channel_indices'][:]

    if 'response_transformation_matrix' in ds[environment].variables:
        control_coordinates = coordinate_array(np.arange(ds[environment]['response_transformation_matrix'].shape[0])+1,0)
        response_transform_comment1 = np.array([f'Unknown :: Transformed Response {i}' for i in np.arange(ds[environment]['response_transformation_matrix'].shape[0])+1],dtype='<U80')
        response_transform_comment2 = np.array([f'Transformed Response {i} :: Transformed Response {i}' for i in np.arange(ds[environment]['response_transformation_matrix'].shape[0])+1],dtype='<U80')
        response_transform_comment3 = np.array([f'Transformed Response {i} :: Transformed Response {i}' for i in np.arange(ds[environment]['response_transformation_matrix'].shape[0])+1],dtype='<U80')
        response_transform_comment4 = np.array([f'Transformed Response {i}' for i in np.arange(ds[environment]['response_transformation_matrix'].shape[0])+1],dtype='<U80')
        response_transform_comment5 = np.array([f'Transformed Response {i}' for i in np.arange(ds[environment]['response_transformation_matrix'].shape[0])+1],dtype='<U80')
        control_indices = np.arange(ds[environment]['response_transformation_matrix'].shape[0])
    else:
        control_coordinates = coordinates[control_indices]

    if 'reference_transformation_matrix' in ds[environment].variables:
        drive_coordinates = coordinate_array(np.arange(ds[environment]['reference_transformation_matrix'].shape[0])+1,0)
        drive_transform_comment1 = np.array([f'Unknown :: Transformed Drive {i}' for i in np.arange(ds[environment]['reference_transformation_matrix'].shape[0])+1],dtype='<U80')
        drive_transform_comment2 = np.array([f'Transformed Drive {i} :: Transformed Drive {i}' for i in np.arange(ds[environment]['reference_transformation_matrix'].shape[0])+1],dtype='<U80')
        drive_transform_comment3 = np.array([f'Transformed Drive {i} :: Transformed Drive {i}' for i in np.arange(ds[environment]['reference_transformation_matrix'].shape[0])+1],dtype='<U80')
        drive_transform_comment4 = np.array([f'Transformed Drive {i}' for i in np.arange(ds[environment]['reference_transformation_matrix'].shape[0])+1],dtype='<U80')
        drive_transform_comment5 = np.array([f'Transformed Drive {i}' for i in np.arange(ds[environment]['reference_transformation_matrix'].shape[0])+1],dtype='<U80')
        drives = np.ones(ds[environment]['reference_transformation_matrix'].shape[0],dtype=bool)
    else:
        drive_coordinates = coordinates[drives]

    # Load the spectral data
    frequency_spacing = ds.sample_rate/ds[environment].sysid_frame_size
    fft_lines = ds[environment].dimensions['sysid_fft_lines'].size
    frequencies = np.arange(fft_lines)*frequency_spacing

    frf_array = np.moveaxis(
        np.array(ds[environment]['frf_data_real'][:]
                 + 1j*ds[environment]['frf_data_imag'][:]),
        0, -1)

    response_cpsd_array = np.moveaxis(
        np.array(ds[environment]['response_cpsd_real'][:]
                 + 1j*ds[environment]['response_cpsd_imag'][:]),
        0, -1)

    drive_cpsd_array = np.moveaxis(
        np.array(ds[environment]['reference_cpsd_real'][:]
                 + 1j*ds[environment]['reference_cpsd_imag'][:]),
        0, -1)

    response_noise_cpsd_array = np.moveaxis(
        np.array(ds[environment]['response_noise_cpsd_real'][:]
                 + 1j*ds[environment]['response_noise_cpsd_imag'][:]),
        0, -1)

    drive_noise_cpsd_array = np.moveaxis(
        np.array(ds[environment]['reference_noise_cpsd_real'][:]
                 + 1j*ds[environment]['reference_noise_cpsd_imag'][:]),
        0, -1)

    coherence_array = np.moveaxis(np.array(ds[environment]['frf_coherence'][:]),
                                  0,-1)

    response_coordinates_cpsd = outer_product(control_coordinates, control_coordinates)
    drive_coordinates_cpsd = outer_product(drive_coordinates, drive_coordinates)
    frf_coordinates = outer_product(control_coordinates,drive_coordinates)
    coherence_coordinates = control_coordinates[:,np.newaxis]

    comment1 = np.char.add(np.char.add(np.array(ds['channels']['channel_type'][:], dtype='<U80'),
                                       np.array(' :: ')),
                           np.array(ds['channels']['unit'][:], dtype='<U80'))
    comment2 = np.char.add(np.char.add(np.array(ds['channels']['physical_device'][:], dtype='<U80'),
                                       np.array(' :: ')),
                           np.array(ds['channels']['physical_channel'][:], dtype='<U80'))
    comment3 = np.char.add(np.char.add(np.array(ds['channels']['feedback_device'][:], dtype='<U80'),
                                       np.array(' :: ')),
                           np.array(ds['channels']['feedback_channel'][:], dtype='<U80'))
    comment4 = np.array(ds['channels']['comment'][:], dtype='<U80')
    comment5 = np.array(ds['channels']['make'][:], dtype='<U80')
    for key in ('model', 'serial_number', 'triax_dof'):
        comment5 = np.char.add(comment5, np.array(' '))
        comment5 = np.char.add(comment5, np.array(ds['channels'][key][:], dtype='<U80'))

    full_comment1 = comment1[environment_channels]
    full_comment2 = comment2[environment_channels]
    full_comment3 = comment3[environment_channels]
    full_comment4 = comment4[environment_channels]
    full_comment5 = comment5[environment_channels]
    
    if 'response_transformation_matrix' in ds[environment].variables:
        comment1 = response_transform_comment1
        comment2 = response_transform_comment2
        comment3 = response_transform_comment3
        comment4 = response_transform_comment4
        comment5 = response_transform_comment5
    else:
        comment1 = full_comment1
        comment2 = full_comment2
        comment3 = full_comment3
        comment4 = full_comment4
        comment5 = full_comment5
    comment1_response_cpsd = np.empty((response_coordinates_cpsd.shape[0], response_coordinates_cpsd.shape[1]), dtype=comment1.dtype)
    comment2_response_cpsd = np.empty((response_coordinates_cpsd.shape[0], response_coordinates_cpsd.shape[1]), dtype=comment1.dtype)
    comment3_response_cpsd = np.empty((response_coordinates_cpsd.shape[0], response_coordinates_cpsd.shape[1]), dtype=comment1.dtype)
    comment4_response_cpsd = np.empty((response_coordinates_cpsd.shape[0], response_coordinates_cpsd.shape[1]), dtype=comment1.dtype)
    comment5_response_cpsd = np.empty((response_coordinates_cpsd.shape[0], response_coordinates_cpsd.shape[1]), dtype=comment1.dtype)
    comment1_coherence = np.empty(response_coordinates_cpsd.shape[0], dtype=comment1.dtype)
    comment2_coherence = np.empty(response_coordinates_cpsd.shape[0], dtype=comment1.dtype)
    comment3_coherence = np.empty(response_coordinates_cpsd.shape[0], dtype=comment1.dtype)
    comment4_coherence = np.empty(response_coordinates_cpsd.shape[0], dtype=comment1.dtype)
    comment5_coherence = np.empty(response_coordinates_cpsd.shape[0], dtype=comment1.dtype)
    for i, idx in enumerate(control_indices):
        comment1_coherence[i] = comment1[idx]
        comment2_coherence[i] = comment2[idx]
        comment3_coherence[i] = comment3[idx]
        comment4_coherence[i] = comment4[idx]
        comment5_coherence[i] = comment5[idx]
        for j, jdx in enumerate(control_indices):
            comment1_response_cpsd[i, j] = comment1[idx] + ' // ' + comment1[jdx]
            comment2_response_cpsd[i, j] = comment2[idx] + ' // ' + comment2[jdx]
            comment3_response_cpsd[i, j] = comment3[idx] + ' // ' + comment3[jdx]
            comment4_response_cpsd[i, j] = comment4[idx] + ' // ' + comment4[jdx]
            comment5_response_cpsd[i, j] = comment5[idx] + ' // ' + comment5[jdx]
        
    if 'reference_transformation_matrix' in ds[environment].variables:
        comment1 = drive_transform_comment1
        comment2 = drive_transform_comment2
        comment3 = drive_transform_comment3
        comment4 = drive_transform_comment4
        comment5 = drive_transform_comment5
    else:
        comment1 = full_comment1
        comment2 = full_comment2
        comment3 = full_comment3
        comment4 = full_comment4
        comment5 = full_comment5
    comment1_drive_cpsd = np.empty((drive_coordinates_cpsd.shape[0], drive_coordinates_cpsd.shape[1]), dtype=comment1.dtype)
    comment2_drive_cpsd = np.empty((drive_coordinates_cpsd.shape[0], drive_coordinates_cpsd.shape[1]), dtype=comment1.dtype)
    comment3_drive_cpsd = np.empty((drive_coordinates_cpsd.shape[0], drive_coordinates_cpsd.shape[1]), dtype=comment1.dtype)
    comment4_drive_cpsd = np.empty((drive_coordinates_cpsd.shape[0], drive_coordinates_cpsd.shape[1]), dtype=comment1.dtype)
    comment5_drive_cpsd = np.empty((drive_coordinates_cpsd.shape[0], drive_coordinates_cpsd.shape[1]), dtype=comment1.dtype)
    drive_indices = np.where(drives)[0]
    for i, idx in enumerate(drive_indices):
        for j, jdx in enumerate(drive_indices):
            comment1_drive_cpsd[i, j] = comment1[idx] + ' // ' + comment1[jdx]
            comment2_drive_cpsd[i, j] = comment2[idx] + ' // ' + comment2[jdx]
            comment3_drive_cpsd[i, j] = comment3[idx] + ' // ' + comment3[jdx]
            comment4_drive_cpsd[i, j] = comment4[idx] + ' // ' + comment4[jdx]
            comment5_drive_cpsd[i, j] = comment5[idx] + ' // ' + comment5[jdx]
                
    if 'response_transformation_matrix' in ds[environment].variables:
        rcomment1 = response_transform_comment1
        rcomment2 = response_transform_comment2
        rcomment3 = response_transform_comment3
        rcomment4 = response_transform_comment4
        rcomment5 = response_transform_comment5
    else:
        rcomment1 = full_comment1
        rcomment2 = full_comment2
        rcomment3 = full_comment3
        rcomment4 = full_comment4
        rcomment5 = full_comment5
    if 'reference_transformation_matrix' in ds[environment].variables:
        dcomment1 = drive_transform_comment1
        dcomment2 = drive_transform_comment2
        dcomment3 = drive_transform_comment3
        dcomment4 = drive_transform_comment4
        dcomment5 = drive_transform_comment5
    else:
        dcomment1 = full_comment1
        dcomment2 = full_comment2
        dcomment3 = full_comment3
        dcomment4 = full_comment4
        dcomment5 = full_comment5
        
    comment1_frf = np.empty((frf_coordinates.shape[0], frf_coordinates.shape[1]), dtype=comment1.dtype)
    comment2_frf = np.empty((frf_coordinates.shape[0], frf_coordinates.shape[1]), dtype=comment1.dtype)
    comment3_frf = np.empty((frf_coordinates.shape[0], frf_coordinates.shape[1]), dtype=comment1.dtype)
    comment4_frf = np.empty((frf_coordinates.shape[0], frf_coordinates.shape[1]), dtype=comment1.dtype)
    comment5_frf = np.empty((frf_coordinates.shape[0], frf_coordinates.shape[1]), dtype=comment1.dtype)
    for i, idx in enumerate(control_indices):
        for j, jdx in enumerate(drive_indices):
            comment1_frf[i, j] = rcomment1[idx] + ' // ' + dcomment1[jdx]
            comment2_frf[i, j] = rcomment2[idx] + ' // ' + dcomment2[jdx]
            comment3_frf[i, j] = rcomment3[idx] + ' // ' + dcomment3[jdx]
            comment4_frf[i, j] = rcomment4[idx] + ' // ' + dcomment4[jdx]
            comment5_frf[i, j] = rcomment5[idx] + ' // ' + dcomment5[jdx]

    # Save the data to SDynpy objects
    response_cpsd = data_array(FunctionTypes.POWER_SPECTRAL_DENSITY,
                               frequencies, response_cpsd_array, response_coordinates_cpsd,
                               comment1_response_cpsd, comment2_response_cpsd, comment3_response_cpsd,
                               comment4_response_cpsd, comment5_response_cpsd)
    response_noise_cpsd = data_array(FunctionTypes.POWER_SPECTRAL_DENSITY,
                               frequencies, response_noise_cpsd_array, response_coordinates_cpsd,
                               comment1_response_cpsd, comment2_response_cpsd, comment3_response_cpsd,
                               comment4_response_cpsd, comment5_response_cpsd)
    drive_cpsd = data_array(FunctionTypes.POWER_SPECTRAL_DENSITY,
                               frequencies, drive_cpsd_array, drive_coordinates_cpsd,
                               comment1_drive_cpsd, comment2_drive_cpsd, comment3_drive_cpsd,
                               comment4_drive_cpsd, comment5_drive_cpsd)
    drive_noise_cpsd = data_array(FunctionTypes.POWER_SPECTRAL_DENSITY,
                               frequencies, drive_noise_cpsd_array, drive_coordinates_cpsd,
                               comment1_drive_cpsd, comment2_drive_cpsd, comment3_drive_cpsd,
                               comment4_drive_cpsd, comment5_drive_cpsd)
    frfs = data_array(FunctionTypes.FREQUENCY_RESPONSE_FUNCTION,
                      frequencies, frf_array, frf_coordinates,
                      comment1_frf,comment2_frf,comment3_frf,comment4_frf,comment5_frf)
    coherence = data_array(FunctionTypes.MULTIPLE_COHERENCE,
                           frequencies,coherence_array,coherence_coordinates,
                           comment1_coherence,comment2_coherence,comment3_coherence,
                           comment4_coherence,comment5_coherence)

    return frfs, response_cpsd, drive_cpsd, response_noise_cpsd, drive_noise_cpsd, coherence

def read_random_spectral_data(file, coordinate_override_column=None):
    if isinstance(file, str):
        ds = nc4.Dataset(file, 'r')
    elif isinstance(file, nc4.Dataset):
        ds = file

    environment = [group for group in ds.groups if not group == 'channels'][0]

    # Get the channels in the group
    if coordinate_override_column is None:
        nodes = [int(''.join(char for char in node if char in '0123456789'))
                 for node in ds['channels']['node_number']]
        directions = np.array(ds['channels']['node_direction'][:], dtype='<U3')
        coordinates = coordinate_array(nodes, directions)
    else:
        coordinates = coordinate_array(string_array=ds['channels'][coordinate_override_column])
    drives = ds['channels']['feedback_device'][:] != ''

    # Cull down to just those in the environment
    environment_index = np.where(ds['environment_names'][:] == environment)[0][0]
    environment_channels = ds['environment_active_channels'][:, environment_index].astype(bool)
    
    drives = drives[environment_channels]
    coordinates = coordinates[environment_channels]

    control_indices = ds[environment]['control_channel_indices'][:]

    if 'response_transformation_matrix' in ds[environment].variables:
        control_coordinates = coordinate_array(np.arange(ds[environment]['response_transformation_matrix'].shape[0])+1,0)
        response_transform_comment1 = np.array([f'Unknown :: Transformed Response {i}' for i in np.arange(ds[environment]['response_transformation_matrix'].shape[0])+1],dtype='<U80')
        response_transform_comment2 = np.array([f'Transformed Response {i} :: Transformed Response {i}' for i in np.arange(ds[environment]['response_transformation_matrix'].shape[0])+1],dtype='<U80')
        response_transform_comment3 = np.array([f'Transformed Response {i} :: Transformed Response {i}' for i in np.arange(ds[environment]['response_transformation_matrix'].shape[0])+1],dtype='<U80')
        response_transform_comment4 = np.array([f'Transformed Response {i}' for i in np.arange(ds[environment]['response_transformation_matrix'].shape[0])+1],dtype='<U80')
        response_transform_comment5 = np.array([f'Transformed Response {i}' for i in np.arange(ds[environment]['response_transformation_matrix'].shape[0])+1],dtype='<U80')
        control_indices = np.arange(ds[environment]['response_transformation_matrix'].shape[0])
    else:
        control_coordinates = coordinates[control_indices]

    if 'reference_transformation_matrix' in ds[environment].variables:
        drive_coordinates = coordinate_array(np.arange(ds[environment]['reference_transformation_matrix'].shape[0])+1,0)
        drive_transform_comment1 = np.array([f'Unknown :: Transformed Drive {i}' for i in np.arange(ds[environment]['reference_transformation_matrix'].shape[0])+1],dtype='<U80')
        drive_transform_comment2 = np.array([f'Transformed Drive {i} :: Transformed Drive {i}' for i in np.arange(ds[environment]['reference_transformation_matrix'].shape[0])+1],dtype='<U80')
        drive_transform_comment3 = np.array([f'Transformed Drive {i} :: Transformed Drive {i}' for i in np.arange(ds[environment]['reference_transformation_matrix'].shape[0])+1],dtype='<U80')
        drive_transform_comment4 = np.array([f'Transformed Drive {i}' for i in np.arange(ds[environment]['reference_transformation_matrix'].shape[0])+1],dtype='<U80')
        drive_transform_comment5 = np.array([f'Transformed Drive {i}' for i in np.arange(ds[environment]['reference_transformation_matrix'].shape[0])+1],dtype='<U80')
        drives = np.ones(ds[environment]['reference_transformation_matrix'].shape[0],dtype=bool)
    else:
        drive_coordinates = coordinates[drives]

    # Load the spectral data
    frequencies = np.array(ds[environment]['specification_frequency_lines'][:])

    spec_cpsd = np.moveaxis(
        np.array(ds[environment]['specification_cpsd_matrix_real'][:]
                 + 1j*ds[environment]['specification_cpsd_matrix_imag'][:]),
        0, -1)

    response_cpsd = np.moveaxis(
        np.array(ds[environment]['response_cpsd_real'][:]
                 + 1j*ds[environment]['response_cpsd_imag'][:]),
        0, -1)

    drive_cpsd = np.moveaxis(
        np.array(ds[environment]['drive_cpsd_real'][:]
                 + 1j*ds[environment]['drive_cpsd_imag'][:]),
        0, -1)

    response_coordinates_cpsd = outer_product(control_coordinates, control_coordinates)
    drive_coordinates_cpsd = outer_product(drive_coordinates, drive_coordinates)

    comment1 = np.char.add(np.char.add(np.array(ds['channels']['channel_type'][:], dtype='<U80'),
                                       np.array(' :: ')),
                           np.array(ds['channels']['unit'][:], dtype='<U80'))
    comment2 = np.char.add(np.char.add(np.array(ds['channels']['physical_device'][:], dtype='<U80'),
                                       np.array(' :: ')),
                           np.array(ds['channels']['physical_channel'][:], dtype='<U80'))
    comment3 = np.char.add(np.char.add(np.array(ds['channels']['feedback_device'][:], dtype='<U80'),
                                       np.array(' :: ')),
                           np.array(ds['channels']['feedback_channel'][:], dtype='<U80'))
    comment4 = np.array(ds['channels']['comment'][:], dtype='<U80')
    comment5 = np.array(ds['channels']['make'][:], dtype='<U80')
    for key in ('model', 'serial_number', 'triax_dof'):
        comment5 = np.char.add(comment5, np.array(' '))
        comment5 = np.char.add(comment5, np.array(ds['channels'][key][:], dtype='<U80'))

    full_comment1 = comment1[environment_channels]
    full_comment2 = comment2[environment_channels]
    full_comment3 = comment3[environment_channels]
    full_comment4 = comment4[environment_channels]
    full_comment5 = comment5[environment_channels]
    
    if 'response_transformation_matrix' in ds[environment].variables:
        comment1 = response_transform_comment1
        comment2 = response_transform_comment2
        comment3 = response_transform_comment3
        comment4 = response_transform_comment4
        comment5 = response_transform_comment5
    else:
        comment1 = full_comment1
        comment2 = full_comment2
        comment3 = full_comment3
        comment4 = full_comment4
        comment5 = full_comment5
    comment1_response_cpsd = np.empty((response_coordinates_cpsd.shape[0], response_coordinates_cpsd.shape[1]), dtype=comment1.dtype)
    comment2_response_cpsd = np.empty((response_coordinates_cpsd.shape[0], response_coordinates_cpsd.shape[1]), dtype=comment1.dtype)
    comment3_response_cpsd = np.empty((response_coordinates_cpsd.shape[0], response_coordinates_cpsd.shape[1]), dtype=comment1.dtype)
    comment4_response_cpsd = np.empty((response_coordinates_cpsd.shape[0], response_coordinates_cpsd.shape[1]), dtype=comment1.dtype)
    comment5_response_cpsd = np.empty((response_coordinates_cpsd.shape[0], response_coordinates_cpsd.shape[1]), dtype=comment1.dtype)
    for i, idx in enumerate(control_indices):
        for j, jdx in enumerate(control_indices):
            comment1_response_cpsd[i, j] = comment1[idx] + ' // ' + comment1[jdx]
            comment2_response_cpsd[i, j] = comment2[idx] + ' // ' + comment2[jdx]
            comment3_response_cpsd[i, j] = comment3[idx] + ' // ' + comment3[jdx]
            comment4_response_cpsd[i, j] = comment4[idx] + ' // ' + comment4[jdx]
            comment5_response_cpsd[i, j] = comment5[idx] + ' // ' + comment5[jdx]
    
    if 'reference_transformation_matrix' in ds[environment].variables:
        comment1 = drive_transform_comment1
        comment2 = drive_transform_comment2
        comment3 = drive_transform_comment3
        comment4 = drive_transform_comment4
        comment5 = drive_transform_comment5
    else:
        comment1 = full_comment1
        comment2 = full_comment2
        comment3 = full_comment3
        comment4 = full_comment4
        comment5 = full_comment5
    comment1_drive_cpsd = np.empty((drive_coordinates_cpsd.shape[0], drive_coordinates_cpsd.shape[1]), dtype=comment1.dtype)
    comment2_drive_cpsd = np.empty((drive_coordinates_cpsd.shape[0], drive_coordinates_cpsd.shape[1]), dtype=comment1.dtype)
    comment3_drive_cpsd = np.empty((drive_coordinates_cpsd.shape[0], drive_coordinates_cpsd.shape[1]), dtype=comment1.dtype)
    comment4_drive_cpsd = np.empty((drive_coordinates_cpsd.shape[0], drive_coordinates_cpsd.shape[1]), dtype=comment1.dtype)
    comment5_drive_cpsd = np.empty((drive_coordinates_cpsd.shape[0], drive_coordinates_cpsd.shape[1]), dtype=comment1.dtype)
    drive_indices = np.where(drives)[0]
    for i, idx in enumerate(drive_indices):
        for j, jdx in enumerate(drive_indices):
            comment1_drive_cpsd[i, j] = comment1[idx] + ' // ' + comment1[jdx]
            comment2_drive_cpsd[i, j] = comment2[idx] + ' // ' + comment2[jdx]
            comment3_drive_cpsd[i, j] = comment3[idx] + ' // ' + comment3[jdx]
            comment4_drive_cpsd[i, j] = comment4[idx] + ' // ' + comment4[jdx]
            comment5_drive_cpsd[i, j] = comment5[idx] + ' // ' + comment5[jdx]
                
    

    # Save the data to SDynpy objects
    response_cpsd = data_array(FunctionTypes.POWER_SPECTRAL_DENSITY,
                               frequencies, response_cpsd, response_coordinates_cpsd,
                               comment1_response_cpsd, comment2_response_cpsd, comment3_response_cpsd,
                               comment4_response_cpsd, comment5_response_cpsd)
    spec_cpsd = data_array(FunctionTypes.POWER_SPECTRAL_DENSITY,
                           frequencies, spec_cpsd, response_coordinates_cpsd,
                           comment1_response_cpsd, comment2_response_cpsd, comment3_response_cpsd,
                           comment4_response_cpsd, comment5_response_cpsd)
    drive_cpsd = data_array(FunctionTypes.POWER_SPECTRAL_DENSITY,
                            frequencies, drive_cpsd, drive_coordinates_cpsd,
                            comment1_drive_cpsd, comment2_drive_cpsd, comment3_drive_cpsd,
                            comment4_drive_cpsd, comment5_drive_cpsd)
    return response_cpsd, spec_cpsd, drive_cpsd


def read_modal_data(file, coordinate_override_column=None, read_only_indices=None):
    if isinstance(file, str):
        ds = nc4.Dataset(file, 'r')
    elif isinstance(file, nc4.Dataset):
        ds = file
    if read_only_indices is None:
        read_only_indices = slice(None)
    # Get parameters
    num_channels = ds.groups['channels'].variables['physical_device'].size
    group_key = [g for g in ds.groups if not g == 'channels'][0]
    group = ds.groups[group_key]
    sample_rate = ds.sample_rate
    samples_per_frame = group.samples_per_frame
    num_averages = group.num_averages
    # Load in the time data
    try:
        output_data = np.array(ds['time_data'][...][read_only_indices]).reshape(num_channels, num_averages, samples_per_frame).transpose(1, 0, 2)
    except ValueError:
        warnings.warn('Number of averages in the time data does not match the number of averages specified in the test settings.  Your test may be incomplete.')
        output_data = np.array(ds['time_data'][...][read_only_indices]).reshape(num_channels, -1, samples_per_frame).transpose(1, 0, 2)
    abscissa = np.arange(samples_per_frame) / sample_rate
    if coordinate_override_column is None:
        nodes = [int(''.join(char for char in node if char in '0123456789'))
                 for node in ds['channels']['node_number'][...][read_only_indices]]
        directions = np.array(ds['channels']['node_direction'][...][read_only_indices], dtype='<U3')
        coordinates = coordinate_array(nodes, directions)[:, np.newaxis]
    else:
        coordinates = coordinate_array(string_array=ds['channels'][coordinate_override_column][read_only_indices])[
            :, np.newaxis]
    array = {name: np.array(variable[:]) for name, variable in ds['channels'].variables.items()}
    channel_table = pd.DataFrame(array)
    comment1 = np.char.add(np.char.add(np.array(ds['channels']['channel_type'][...][read_only_indices], dtype='<U80'),
                                       np.array(' :: ')),
                           np.array(ds['channels']['unit'][...][read_only_indices], dtype='<U80'))
    comment2 = np.char.add(np.char.add(np.array(ds['channels']['physical_device'][...][read_only_indices], dtype='<U80'),
                                       np.array(' :: ')),
                           np.array(ds['channels']['physical_channel'][...][read_only_indices], dtype='<U80'))
    comment3 = np.char.add(np.char.add(np.array(ds['channels']['feedback_device'][...][read_only_indices], dtype='<U80'),
                                       np.array(' :: ')),
                           np.array(ds['channels']['feedback_channel'][...][read_only_indices], dtype='<U80'))
    comment4 = np.array(ds['channels']['comment'][...][read_only_indices], dtype='<U80')
    comment5 = np.array(ds['channels']['make'][...][read_only_indices], dtype='<U80')
    for key in ('model', 'serial_number', 'triax_dof'):
        comment5 = np.char.add(comment5, np.array(' '))
        comment5 = np.char.add(comment5, np.array(ds['channels'][key][...][read_only_indices], dtype='<U80'))
    time_data = data_array(FunctionTypes.TIME_RESPONSE,
                           abscissa,
                           output_data,
                           coordinates,
                           comment1,
                           comment2,
                           comment3,
                           comment4,
                           comment5)
    # Response and Reference Indices
    kept_indices = np.arange(num_channels)[read_only_indices]
    reference_indices = np.array(group.variables['reference_channel_indices'][:])
    response_indices = np.array(group.variables['response_channel_indices'][:])
    keep_response_indices = np.array([i for i, index in enumerate(response_indices) if index in kept_indices])
    keep_reference_indices = np.array([i for i, index in enumerate(reference_indices) if index in kept_indices])
    frequency_lines = np.arange(group.dimensions['fft_lines'].size)*sample_rate/samples_per_frame
    coherence_data = np.array(group['coherence'][:, keep_response_indices]).T
    comment1 = np.char.add(np.char.add(np.array(ds['channels']['channel_type'][...][response_indices[keep_response_indices]], dtype='<U80'),
                                       np.array(' :: ')),
                           np.array(ds['channels']['unit'][...][response_indices[keep_response_indices]], dtype='<U80'))
    comment2 = np.char.add(np.char.add(np.array(ds['channels']['physical_device'][...][response_indices[keep_response_indices]], dtype='<U80'),
                                       np.array(' :: ')),
                           np.array(ds['channels']['physical_channel'][...][response_indices[keep_response_indices]], dtype='<U80'))
    comment3 = np.char.add(np.char.add(np.array(ds['channels']['feedback_device'][...][response_indices[keep_response_indices]], dtype='<U80'),
                                       np.array(' :: ')),
                           np.array(ds['channels']['feedback_channel'][...][response_indices[keep_response_indices]], dtype='<U80'))
    comment4 = np.array(ds['channels']['comment'][...][response_indices[keep_response_indices]], dtype='<U80')
    comment5 = np.array(ds['channels']['make'][...][response_indices[keep_response_indices]], dtype='<U80')
    for key in ('model', 'serial_number', 'triax_dof'):
        comment5 = np.char.add(comment5, np.array(' '))
        comment5 = np.char.add(comment5, np.array(ds['channels'][key][...][response_indices[keep_response_indices]], dtype='<U80'))
    coherence_data = data_array(FunctionTypes.MULTIPLE_COHERENCE,
                                frequency_lines,
                                coherence_data,
                                coordinates[response_indices[keep_response_indices]],
                                comment1,
                                comment2,
                                comment3,
                                comment4,
                                comment5)
    # Frequency Response Functions
    frf_data = np.moveaxis(np.array(group['frf_data_real'])[:, keep_response_indices[:, np.newaxis], keep_reference_indices]
                           + np.array(group['frf_data_imag'])[:, keep_response_indices[:, np.newaxis], keep_reference_indices]*1j, 0, -1)
    frf_coordinate = outer_product(coordinates[response_indices[keep_response_indices], 0],
                                   coordinates[reference_indices[keep_reference_indices], 0])
    # print(response_indices[keep_response_indices])
    # print(reference_indices[keep_reference_indices])
    response_comment1 = np.char.add(np.char.add(np.array(ds['channels']['channel_type'][...][response_indices[keep_response_indices]], dtype='<U80'),
                                                np.array(' :: ')),
                                    np.array(ds['channels']['unit'][...][response_indices[keep_response_indices]], dtype='<U80'))
    response_comment2 = np.char.add(np.char.add(np.array(ds['channels']['physical_device'][...][response_indices[keep_response_indices]], dtype='<U80'),
                                                np.array(' :: ')),
                                    np.array(ds['channels']['physical_channel'][...][response_indices[keep_response_indices]], dtype='<U80'))
    response_comment3 = np.char.add(np.char.add(np.array(ds['channels']['feedback_device'][...][response_indices[keep_response_indices]], dtype='<U80'),
                                                np.array(' :: ')),
                                    np.array(ds['channels']['feedback_channel'][...][response_indices[keep_response_indices]], dtype='<U80'))
    response_comment4 = np.array(ds['channels']['comment'][...][response_indices[keep_response_indices]], dtype='<U80')
    response_comment5 = np.array(ds['channels']['make'][...][response_indices[keep_response_indices]], dtype='<U80')
    for key in ('model', 'serial_number', 'triax_dof'):
        response_comment5 = np.char.add(response_comment5, np.array(' '))
        response_comment5 = np.char.add(response_comment5, np.array(ds['channels'][key][...][response_indices[keep_response_indices]], dtype='<U80'))
    reference_comment1 = np.char.add(np.char.add(np.array(ds['channels']['channel_type'][...][reference_indices[keep_reference_indices]], dtype='<U80'),
                                                 np.array(' :: ')),
                                     np.array(ds['channels']['unit'][...][reference_indices[keep_reference_indices]], dtype='<U80'))
    reference_comment2 = np.char.add(np.char.add(np.array(ds['channels']['physical_device'][...][reference_indices[keep_reference_indices]], dtype='<U80'),
                                                 np.array(' :: ')),
                                     np.array(ds['channels']['physical_channel'][...][reference_indices[keep_reference_indices]], dtype='<U80'))
    reference_comment3 = np.char.add(np.char.add(np.array(ds['channels']['feedback_device'][...][reference_indices[keep_reference_indices]], dtype='<U80'),
                                                 np.array(' :: ')),
                                     np.array(ds['channels']['feedback_channel'][...][reference_indices[keep_reference_indices]], dtype='<U80'))
    reference_comment4 = np.array(ds['channels']['comment'][...][reference_indices[keep_reference_indices]], dtype='<U80')
    reference_comment5 = np.array(ds['channels']['make'][...][reference_indices[keep_reference_indices]], dtype='<U80')
    for key in ('model', 'serial_number', 'triax_dof'):
        reference_comment5 = np.char.add(reference_comment5, np.array(' '))
        reference_comment5 = np.char.add(reference_comment5, np.array(ds['channels'][key][...][reference_indices[keep_reference_indices]], dtype='<U80'))
    response_comment1, reference_comment1 = np.broadcast_arrays(response_comment1[:, np.newaxis], reference_comment1)
    comment1 = np.char.add(np.char.add(response_comment1, np.array(' / ')), reference_comment1)
    response_comment2, reference_comment2 = np.broadcast_arrays(response_comment2[:, np.newaxis], reference_comment2)
    comment2 = np.char.add(np.char.add(response_comment2, np.array(' / ')), reference_comment2)
    response_comment3, reference_comment3 = np.broadcast_arrays(response_comment3[:, np.newaxis], reference_comment3)
    comment3 = np.char.add(np.char.add(response_comment3, np.array(' / ')), reference_comment3)
    response_comment4, reference_comment4 = np.broadcast_arrays(response_comment4[:, np.newaxis], reference_comment4)
    comment4 = np.char.add(np.char.add(response_comment4, np.array(' / ')), reference_comment4)
    response_comment5, reference_comment5 = np.broadcast_arrays(response_comment5[:, np.newaxis], reference_comment5)
    comment5 = np.char.add(np.char.add(response_comment5, np.array(' / ')), reference_comment5)
    frf_data = data_array(FunctionTypes.FREQUENCY_RESPONSE_FUNCTION,
                          frequency_lines,
                          frf_data,
                          frf_coordinate,
                          comment1,
                          comment2,
                          comment3,
                          comment4,
                          comment5)
    return time_data, frf_data, coherence_data, channel_table


def read_transient_control_data(file, coordinate_override_column=None):
    if isinstance(file, str):
        ds = nc4.Dataset(file, 'r')
    elif isinstance(file, nc4.Dataset):
        ds = file
    coordinate_override_column = None

    environment = [group for group in ds.groups if not group == 'channels'][0]

    # Get the channels in the group
    if coordinate_override_column is None:
        nodes = [int(''.join(char for char in node if char in '0123456789'))
                 for node in ds['channels']['node_number']]
        directions = np.array(ds['channels']['node_direction'][:], dtype='<U3')
        coordinates = coordinate_array(nodes, directions)
    else:
        coordinates = coordinate_array(string_array=ds['channels'][coordinate_override_column])
    drives = ds['channels']['feedback_device'][:] != ''

    # Cull down to just those in the environment
    environment_index = np.where(ds['environment_names'][:] == environment)[0][0]
    environment_channels = ds['environment_active_channels'][:, environment_index].astype(bool)

    drives = drives[environment_channels]
    coordinates = coordinates[environment_channels]

    control_indices = ds[environment]['control_channel_indices'][:]

    if 'response_transformation_matrix' in ds[environment].variables:
        control_coordinates = coordinate_array(np.arange(ds[environment]['response_transformation_matrix'].shape[0])+1,0)
        response_transform_comment1 = np.array([f'Unknown :: Transformed Response {i}' for i in np.arange(ds[environment]['response_transformation_matrix'].shape[0])+1],dtype='<U80')
        response_transform_comment2 = np.array([f'Transformed Response {i} :: Transformed Response {i}' for i in np.arange(ds[environment]['response_transformation_matrix'].shape[0])+1],dtype='<U80')
        response_transform_comment3 = np.array([f'Transformed Response {i} :: Transformed Response {i}' for i in np.arange(ds[environment]['response_transformation_matrix'].shape[0])+1],dtype='<U80')
        response_transform_comment4 = np.array([f'Transformed Response {i}' for i in np.arange(ds[environment]['response_transformation_matrix'].shape[0])+1],dtype='<U80')
        response_transform_comment5 = np.array([f'Transformed Response {i}' for i in np.arange(ds[environment]['response_transformation_matrix'].shape[0])+1],dtype='<U80')
        control_indices = np.arange(ds[environment]['response_transformation_matrix'].shape[0])
    else:
        control_coordinates = coordinates[control_indices]

    if 'reference_transformation_matrix' in ds[environment].variables:
        drive_coordinates = coordinate_array(np.arange(ds[environment]['reference_transformation_matrix'].shape[0])+1,0)
        drive_transform_comment1 = np.array([f'Unknown :: Transformed Drive {i}' for i in np.arange(ds[environment]['reference_transformation_matrix'].shape[0])+1],dtype='<U80')
        drive_transform_comment2 = np.array([f'Transformed Drive {i} :: Transformed Drive {i}' for i in np.arange(ds[environment]['reference_transformation_matrix'].shape[0])+1],dtype='<U80')
        drive_transform_comment3 = np.array([f'Transformed Drive {i} :: Transformed Drive {i}' for i in np.arange(ds[environment]['reference_transformation_matrix'].shape[0])+1],dtype='<U80')
        drive_transform_comment4 = np.array([f'Transformed Drive {i}' for i in np.arange(ds[environment]['reference_transformation_matrix'].shape[0])+1],dtype='<U80')
        drive_transform_comment5 = np.array([f'Transformed Drive {i}' for i in np.arange(ds[environment]['reference_transformation_matrix'].shape[0])+1],dtype='<U80')
        drives = np.ones(ds[environment]['reference_transformation_matrix'].shape[0],dtype=bool)
    else:
        drive_coordinates = coordinates[drives]

    # Load the time data
    timesteps = np.arange(ds[environment].dimensions['signal_samples'].size)/ds.sample_rate

    spec_signal = np.array(ds[environment]['control_signal'][...])

    response_signal = np.array(ds[environment]['control_response'][...])

    drive_signal = np.array(ds[environment]['control_drives'][...])

    response_coordinates = control_coordinates[:, np.newaxis]
    drive_coordinates = drive_coordinates[:, np.newaxis]

    comment1 = np.char.add(np.char.add(np.array(ds['channels']['channel_type'][:], dtype='<U80'),
                                       np.array(' :: ')),
                           np.array(ds['channels']['unit'][:], dtype='<U80'))
    comment2 = np.char.add(np.char.add(np.array(ds['channels']['physical_device'][:], dtype='<U80'),
                                       np.array(' :: ')),
                           np.array(ds['channels']['physical_channel'][:], dtype='<U80'))
    comment3 = np.char.add(np.char.add(np.array(ds['channels']['feedback_device'][:], dtype='<U80'),
                                       np.array(' :: ')),
                           np.array(ds['channels']['feedback_channel'][:], dtype='<U80'))
    comment4 = np.array(ds['channels']['comment'][:], dtype='<U80')
    comment5 = np.array(ds['channels']['make'][:], dtype='<U80')
    for key in ('model', 'serial_number', 'triax_dof'):
        comment5 = np.char.add(comment5, np.array(' '))
        comment5 = np.char.add(comment5, np.array(ds['channels'][key][:], dtype='<U80'))

    full_comment1 = comment1[environment_channels]
    full_comment2 = comment2[environment_channels]
    full_comment3 = comment3[environment_channels]
    full_comment4 = comment4[environment_channels]
    full_comment5 = comment5[environment_channels]

    if 'response_transformation_matrix' in ds[environment].variables:
        comment1 = response_transform_comment1
        comment2 = response_transform_comment2
        comment3 = response_transform_comment3
        comment4 = response_transform_comment4
        comment5 = response_transform_comment5
    else:
        comment1 = full_comment1[control_indices]
        comment2 = full_comment2[control_indices]
        comment3 = full_comment3[control_indices]
        comment4 = full_comment4[control_indices]
        comment5 = full_comment5[control_indices]

    # Save the data to SDynpy objects
    response_signal = data_array(FunctionTypes.TIME_RESPONSE,
                                 timesteps, response_signal, response_coordinates,
                                 comment1, comment2, comment3,
                                 comment4, comment5)
    spec_signal = data_array(FunctionTypes.TIME_RESPONSE,
                             timesteps, spec_signal, response_coordinates,
                             comment1, comment2, comment3,
                             comment4, comment5)
    
    if 'reference_transformation_matrix' in ds[environment].variables:
        comment1 = drive_transform_comment1
        comment2 = drive_transform_comment2
        comment3 = drive_transform_comment3
        comment4 = drive_transform_comment4
        comment5 = drive_transform_comment5
    else:
        comment1 = full_comment1[drives]
        comment2 = full_comment2[drives]
        comment3 = full_comment3[drives]
        comment4 = full_comment4[drives]
        comment5 = full_comment5[drives]

    drive_signal = data_array(FunctionTypes.TIME_RESPONSE,
                              timesteps, drive_signal, drive_coordinates,
                              comment1, comment2, comment3,
                              comment4, comment5)
    
    return response_signal, spec_signal, drive_signal


def create_synthetic_test(spreadsheet_file_name: str,
                          system_filename: str, system: System,
                          excitation_coordinates: CoordinateArray,
                          response_coordinates: CoordinateArray,
                          rattlesnake_directory: str,
                          displacement_derivative=2,
                          sample_rate: int = None,
                          time_per_read: float = None,
                          time_per_write: float = None,
                          integration_oversample: int = 10,
                          environments: list = [],
                          channel_comment_data: list = None,
                          channel_serial_number_data: list = None,
                          channel_triax_dof_data: list = None,
                          channel_engineering_unit_data: list = None,
                          channel_warning_level_data: list = None,
                          channel_abort_level_data: list = None,
                          channel_active_in_environment_data: dict = None
                          ):
    system.save(system_filename)
    # Load in Rattlesnake to create a template for the test
    sys.path.insert(0, rattlesnake_directory)
    import components as rs
    environment_data = []
    for environment_type, environment_name in environments:
        # Find the identifier
        environment_type = rs.environments.ControlTypes[environment_type.upper()]
        environment_data.append((environment_type, environment_name))
    rs.ui_utilities.save_combined_environments_profile_template(spreadsheet_file_name, environment_data)
    sys.path.pop(0)
    # Populate the channel table
    workbook = opxl.load_workbook(spreadsheet_file_name)
    worksheet = workbook.get_sheet_by_name('Channel Table')
    index = 3
    for i, channel in enumerate(response_coordinates):
        worksheet.cell(index, 1, i+1)
        worksheet.cell(index, 2, channel.node)
        worksheet.cell(index, 3, _string_map[channel.direction])
        worksheet.cell(index, 12, 'Virtual')
        worksheet.cell(index, 14, 'Accel')
        index += 1
    for i, channel in enumerate(excitation_coordinates):
        worksheet.cell(index, 1, len(response_coordinates)+i+1)
        worksheet.cell(index, 2, channel.node)
        worksheet.cell(index, 3, _string_map[channel.direction])
        worksheet.cell(index, 12, 'Virtual')
        worksheet.cell(index, 14, 'Force')
        worksheet.cell(index, 20, 'Shaker')
        index += 1
    # Go through the various channel table data that could have been optionally
    # provided
    for column, data in [(4, channel_comment_data),
                         (5, channel_serial_number_data),
                         (6, channel_triax_dof_data),
                         (8, channel_engineering_unit_data),
                         (22, channel_warning_level_data),
                         (23, channel_abort_level_data)]:
        if data is None:
            continue
        for row_index, value in enumerate(data):
            worksheet.cell(3+row_index, column, value)
    # Now fill out the environment table
    if channel_active_in_environment_data is not None:
        for environment_index, (environment_type, environment_name) in enumerate(environment_data):
            for row_index, value in enumerate(channel_active_in_environment_data[environment_name]):
                if value:
                    worksheet.cell(3+row_index, 24+environment_index, 'X')
    else:
        for environment_index, (environment_type, environment_name) in enumerate(environment_data):
            for row_index in range(response_coordinates.size + excitation_coordinates.size):
                worksheet.cell(3+row_index, 24+environment_index, 'X')
    worksheet = workbook.get_sheet_by_name('Hardware')
    worksheet.cell(1, 2, 6)
    worksheet.cell(2, 2, os.path.abspath(system_filename))
    if sample_rate is not None:
        worksheet.cell(3, 2, sample_rate)
    if time_per_read is not None:
        worksheet.cell(4, 2, time_per_read)
    if time_per_write is not None:
        worksheet.cell(5, 2, time_per_write)
    worksheet.cell(6, 2, 1)
    worksheet.cell(7, 2, integration_oversample)
    workbook.save(spreadsheet_file_name)

def read_sine_control_data(control_file, 
                           read_quantities = 'control_response_signals_combined',
                           excitation_dofs = None, control_dofs = None):
    concatenated_keys = ['control_response_signals_combined',
                         'control_response_amplitudes',
                         'control_response_phases',
                         'control_drive_modifications']
    unconcatenated_keys = ['control_response_frequencies',
                           'control_response_arguments',
                           'control_target_phases',
                           'control_target_amplitudes']
    dimension_labels = {}
    dimension_labels['control_response_signals_combined'] = ('response','timestep')
    dimension_labels['control_response_amplitudes'] = ('tone','response','timestep')
    dimension_labels['control_response_phases'] = ('tone','response','timestep')
    dimension_labels['control_drive_modifications'] = ('tone','excitation','block_num')
    dimension_labels['achieved_excitation_signals_combined'] = ('excitation','timestep')
    dimension_labels['achieved_excitation_signals'] = ('tone','excitation','timestep')
    dimension_labels['control_response_frequencies'] = ('tone','timestep')
    dimension_labels['control_response_arguments'] = ('tone','timestep')
    dimension_labels['control_target_amplitudes'] = ('tone','response','timestep')
    dimension_labels['control_target_phases'] = ('tone','response','timestep')
    if isinstance(control_file,str):
        control_file = np.load(control_file)
    sample_rate = control_file['sample_rate']
    if isinstance(read_quantities,str):
        read_quantities = [read_quantities]
        return_single = True
    else:
        return_single = False
    return_data = []
    for read_quantity in read_quantities:
        try:
            dimension_label = dimension_labels[read_quantity]
        except KeyError:
            raise ValueError(f'{read_quantity} is not a valid quantity to read.  read_quantity must be one of {concatenated_keys+unconcatenated_keys}.')
        # Extract the data and concatenate if necessary
        if read_quantity in concatenated_keys:
            data = []
            for key in control_file:
                if read_quantity == '_'.join(key.split("_")[:-1]):
                    this_data = control_file[key]
                    while this_data.ndim < len(dimension_label):
                        this_data = this_data[...,np.newaxis]
                    data.append(this_data)
            data = np.concatenate(data,axis=-1)
        elif read_quantity in unconcatenated_keys:
            data = control_file[read_quantity]
        else:
            raise ValueError(f'{read_quantity} is not a valid quantity to read.  read_quantity must be one of {concatenated_keys+unconcatenated_keys}.')
        # Set up the abscissa
        if dimension_label[-1] == 'timestep':
            abscissa = np.arange(data.shape[-1])/sample_rate
        elif dimension_label[-1] == 'block_num':
            abscissa = np.arange(data.shape[-1])
        else:
            raise ValueError(f"{dimension_label[-1]} is an invalid entry.  How did you get here?")
        # Set up degrees of freedom
        if dimension_label[-2] == 'response':
            if control_dofs is None:
                dofs = coordinate_array(np.arange(data.shape[-2])+1,0)
            else:
                dofs = control_dofs
        elif dimension_label[-2] == 'excitation':
            if excitation_dofs is None:
                dofs = coordinate_array(np.arange(data.shape[-2])+1,0)
            else:
                dofs = excitation_dofs
        elif dimension_label[-2] == 'tone':
            dofs = coordinate_array(np.arange(data.shape[-2])+1,0)
        else:
            raise ValueError(f"{dimension_label[-2]} is an invalid entry.  How did you get here?")
        if any([dimension == 'tone' for dimension in dimension_label]):
            comment1 = control_file['names'].reshape(*[-1 if dimension == 'tone' else 1 for dimension in dimension_label][:-1])
        else:
            comment1 = ''
        # Construct the TimeHistoryArray
        return_data.append(data_array(FunctionTypes.TIME_RESPONSE,
                                      abscissa,
                                      data,
                                      dofs,
                                      comment1
                                      ))
    if return_single:
        return_data = return_data[0]
    return return_data
        